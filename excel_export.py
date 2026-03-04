"""
excel_export.py
📊 Exportación profesional a Excel con formato financiero estándar.
Genera un workbook con múltiples hojas:
  1. Resumen / KPIs
  2. Datos del Portafolio
  3. Transacciones
  4. Backtesting
  5. Predicciones
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime


# ─────────────────────────────────────────────────────────────────────────────
# Paleta de colores (estándar financiero)
# ─────────────────────────────────────────────────────────────────────────────
DARK_BLUE   = '1F3864'
MID_BLUE    = '2E75B6'
LIGHT_BLUE  = 'BDD7EE'
DARK_GRAY   = '404040'
LIGHT_GRAY  = 'F2F2F2'
GREEN       = '70AD47'
RED         = 'FF0000'
ORANGE      = 'ED7D31'
YELLOW_BG   = 'FFFF00'   # assumptions
WHITE       = 'FFFFFF'
BLACK       = '000000'
INPUT_BLUE  = '0000FF'   # hardcoded inputs (industry standard)
FORMULA_BLK = '000000'   # formulas (industry standard)
LINK_GREEN  = '008000'   # cross-sheet links (industry standard)


def _border(style='thin'):
    s = Side(style=style, color=DARK_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font(size=11, bold=True, color=WHITE):
    return Font(name='Arial', size=size, bold=bold, color=color)

def _body_font(size=10, bold=False, color=BLACK):
    return Font(name='Arial', size=size, bold=bold, color=color)

def _fill(hex_color: str):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _right():
    return Alignment(horizontal='right', vertical='center')

def _auto_col_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 1: Dashboard / KPIs
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_kpi(wb: Workbook, kpis: dict):
    ws = wb.active
    ws.title = 'Dashboard KPIs'
    ws.sheet_view.showGridLines = False

    # Título principal
    ws.merge_cells('B2:G3')
    title_cell = ws['B2']
    title_cell.value     = '📊 Financial Analytics Dashboard'
    title_cell.font      = Font(name='Arial', size=18, bold=True, color=WHITE)
    title_cell.fill      = _fill(DARK_BLUE)
    title_cell.alignment = _center()

    # Subtítulo con fecha
    ws.merge_cells('B4:G4')
    ws['B4'].value     = f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['B4'].font      = _body_font(size=9, color=DARK_GRAY)
    ws['B4'].alignment = _center()
    ws['B4'].fill      = _fill(LIGHT_GRAY)

    # Headers de KPIs
    ws['B6'].value = 'Indicador'
    ws['C6'].value = 'Valor'
    ws['D6'].value = 'Descripción'

    for col in ['B', 'C', 'D']:
        c = ws[f'{col}6']
        c.font      = _header_font()
        c.fill      = _fill(MID_BLUE)
        c.alignment = _center()
        c.border    = _border()

    # Descripciones de KPIs
    descriptions = {
        'Total Ingresos':     'Suma de todos los ingresos del período',
        'Total Egresos':      'Suma de todos los egresos del período',
        'Balance Neto':       'Ingresos menos egresos totales',
        'Ratio de Ahorro %':  'Porcentaje del ingreso que se ahorra',
        'Num. Transacciones': 'Total de movimientos registrados',
        'Ticket Promedio':    'Monto promedio por transacción',
    }

    for i, (label, value) in enumerate(kpis.items(), start=7):
        row_fill = _fill(WHITE) if i % 2 == 0 else _fill(LIGHT_GRAY)

        ws[f'B{i}'].value     = label
        ws[f'B{i}'].font      = _body_font(bold=True)
        ws[f'B{i}'].fill      = row_fill
        ws[f'B{i}'].border    = _border()

        ws[f'C{i}'].value     = value
        ws[f'C{i}'].font      = Font(name='Arial', size=10, color=INPUT_BLUE, bold=True)
        ws[f'C{i}'].fill      = row_fill
        ws[f'C{i}'].alignment = _right()
        ws[f'C{i}'].border    = _border()

        # Formato de número
        if isinstance(value, float) and abs(value) > 10:
            ws[f'C{i}'].number_format = '$#,##0.00;($#,##0.00);"-"'
        elif 'Ratio' in label or '%' in label:
            ws[f'C{i}'].number_format = '0.0"%"'

        ws[f'D{i}'].value     = descriptions.get(label, '')
        ws[f'D{i}'].font      = _body_font(color=DARK_GRAY)
        ws[f'D{i}'].fill      = row_fill
        ws[f'D{i}'].border    = _border()

    # Leyenda de colores
    ws['B15'].value = '📌 Leyenda de colores:'
    ws['B15'].font  = _body_font(bold=True)
    for row_data in [
        (16, INPUT_BLUE, 'Azul', 'Inputs / valores ingresados manualmente'),
        (17, FORMULA_BLK, 'Negro', 'Fórmulas calculadas'),
        (18, LINK_GREEN, 'Verde', 'Vínculos a otras hojas del libro'),
    ]:
        r, color, name, desc = row_data
        ws[f'B{r}'].value = name
        ws[f'B{r}'].font  = Font(name='Arial', size=9, color=color)
        ws[f'C{r}'].value = desc
        ws[f'C{r}'].font  = _body_font(size=9, color=DARK_GRAY)

    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 45
    ws.row_dimensions[2].height = 30


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 2: Datos del portafolio
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_portfolio(wb: Workbook, portfolio: dict):
    ws = wb.create_sheet('Portafolio')
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'].value     = '📈 Datos del Portafolio de Acciones'
    ws['A1'].font      = _header_font(size=14)
    ws['A1'].fill      = _fill(DARK_BLUE)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    col_offset = 1
    for ticker, df in portfolio.items():
        sample = df.tail(60).reset_index()   # últimos 60 días hábiles
        headers = ['Fecha', 'Apertura', 'Máximo', 'Mínimo', 'Cierre', 'Volumen', 'Retorno %']

        # Header de ticker
        end_col = col_offset + len(headers) - 1
        ws.merge_cells(start_row=3, start_column=col_offset,
                       end_row=3,   end_column=end_col)
        ticker_cell = ws.cell(row=3, column=col_offset, value=f'▶ {ticker}')
        ticker_cell.font      = _header_font(size=11)
        ticker_cell.fill      = _fill(MID_BLUE)
        ticker_cell.alignment = _center()

        # Sub-headers
        for j, h in enumerate(headers):
            c = ws.cell(row=4, column=col_offset + j, value=h)
            c.font      = _header_font(size=9)
            c.fill      = _fill(DARK_GRAY)
            c.alignment = _center()
            c.border    = _border()

        # Datos
        prev_close = None
        for i, row in enumerate(sample.itertuples(), start=5):
            ret = ((row.Close / prev_close) - 1) * 100 if prev_close else 0
            prev_close = row.Close
            row_fill = _fill(WHITE) if i % 2 == 0 else _fill(LIGHT_GRAY)

            values = [row.Date.strftime('%d/%m/%Y') if hasattr(row.Date, 'strftime')
                      else str(row.Date),
                      round(row.Open, 2), round(row.High, 2),
                      round(row.Low, 2),  round(row.Close, 2),
                      row.Volume, round(ret, 2)]

            for j, val in enumerate(values):
                c = ws.cell(row=i, column=col_offset + j, value=val)
                c.font   = _body_font(size=9)
                c.fill   = row_fill
                c.border = _border()

                # Formato monetario
                if j in [1, 2, 3, 4]:
                    c.number_format = '$#,##0.00'
                elif j == 5:
                    c.number_format = '#,##0'
                elif j == 6:
                    c.number_format = '0.00"%"'
                    if isinstance(val, (int, float)) and val < 0:
                        c.font = Font(name='Arial', size=9, color=RED)
                    elif isinstance(val, (int, float)) and val > 0:
                        c.font = Font(name='Arial', size=9, color=GREEN)

        col_offset += len(headers) + 2   # espacio entre tablas

    _auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 3: Transacciones
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_transactions(wb: Workbook, transactions: pd.DataFrame):
    ws = wb.create_sheet('Transacciones')
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'].value     = '💳 Registro de Transacciones Financieras'
    ws['A1'].font      = _header_font(size=14)
    ws['A1'].fill      = _fill(DARK_BLUE)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    # Headers
    headers = ['Fecha', 'Categoría', 'Tipo', 'Banco', 'Monto ($)', 'Balance Acumulado ($)', 'Notas']
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = _header_font(); c.fill = _fill(MID_BLUE)
        c.alignment = _center(); c.border = _border()

    # Datos
    for i, row in enumerate(transactions.itertuples(), start=4):
        row_fill = _fill(WHITE) if i % 2 == 0 else _fill(LIGHT_GRAY)
        fecha    = row.Fecha.strftime('%d/%m/%Y') if hasattr(row.Fecha, 'strftime') else str(row.Fecha)

        vals = [fecha, row.Categoría, row.Tipo, row.Banco,
                row.Monto, row.Balance_Acumulado, '']

        for j, val in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = _body_font(size=9); c.fill = row_fill; c.border = _border()

            if j == 5:   # Monto
                c.number_format = '$#,##0.00;($#,##0.00);"-"'
                if isinstance(val, (int,float)) and val < 0:
                    c.font = Font(name='Arial', size=9, color=RED)
                else:
                    c.font = Font(name='Arial', size=9, color=GREEN)
            elif j == 6:  # Balance
                c.number_format = '$#,##0.00;($#,##0.00);"-"'

    # Fila de totales con fórmulas Excel
    last = 3 + len(transactions)
    total_row = last + 2

    ws.cell(row=total_row, column=4, value='TOTALES').font = _header_font(color=WHITE)
    ws.cell(row=total_row, column=4).fill = _fill(DARK_BLUE)

    for col, formula in [(5, f'=SUM(E4:E{last})'), (6, f'=E{last}')]:
        c = ws.cell(row=total_row, column=col, value=formula)
        c.font         = Font(name='Arial', size=10, bold=True, color=FORMULA_BLK)
        c.fill         = _fill(LIGHT_BLUE)
        c.number_format = '$#,##0.00;($#,##0.00);"-"'
        c.border       = _border()

    _auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 4: Backtesting
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_backtest(wb: Workbook, bt_results: dict):
    ws = wb.create_sheet('Backtesting')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:H1')
    ws['A1'].value     = '🔬 Resultados de Backtesting de Estrategias'
    ws['A1'].font      = _header_font(size=14)
    ws['A1'].fill      = _fill(DARK_BLUE)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    # Tabla de métricas comparativas
    ws['A3'].value = 'Comparativa de Estrategias'
    ws['A3'].font  = _header_font()
    ws['A3'].fill  = _fill(MID_BLUE)

    metric_headers = ['Estrategia', 'Capital Inicial', 'Capital Final',
                      'Retorno Total (%)', 'Retorno Anual (%)',
                      'Sharpe Ratio', 'Max Drawdown (%)', 'Num. Operaciones']

    for j, h in enumerate(metric_headers, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = _header_font(size=9); c.fill = _fill(DARK_GRAY)
        c.alignment = _center(); c.border = _border()

    for i, (name, result) in enumerate(bt_results.items(), start=5):
        m = result.attrs.get('metrics', {})
        row_fill = _fill(WHITE) if i % 2 == 0 else _fill(LIGHT_GRAY)
        vals = [m.get(k, '') for k in metric_headers]

        for j, val in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = _body_font(); c.fill = row_fill; c.border = _border()
            c.alignment = _center()

            if j in [2, 3]:
                c.number_format = '$#,##0.00'
            elif j in [4, 5, 6, 7]:
                c.number_format = '0.00'

            # Color rojo/verde en retorno
            if j == 4 and isinstance(val, (int,float)):
                c.font = Font(name='Arial', size=10,
                              color=GREEN if val >= 0 else RED)

    # Datos de equity curves
    row_offset = 8 + len(bt_results)
    ws.cell(row=row_offset, column=1, value='Evolución del Portafolio').font = _header_font()
    ws.cell(row=row_offset, column=1).fill = _fill(MID_BLUE)
    row_offset += 1

    header_row = row_offset
    ws.cell(row=header_row, column=1, value='Fecha').font = _header_font(size=9)
    ws.cell(row=header_row, column=1).fill = _fill(DARK_GRAY)
    ws.cell(row=header_row, column=1).border = _border()

    strategy_cols = {}
    for j, name in enumerate(bt_results.keys(), start=2):
        c = ws.cell(row=header_row, column=j, value=name)
        c.font = _header_font(size=9); c.fill = _fill(DARK_GRAY); c.border = _border()
        strategy_cols[name] = j

    # Muestrear cada 5 días para no sobrecargar
    sample_idx = None
    for name, result in bt_results.items():
        sample_idx = result.index[::5]
        break

    for i, date in enumerate(sample_idx, start=header_row + 1):
        date_str = date.strftime('%d/%m/%Y') if hasattr(date, 'strftime') else str(date)
        ws.cell(row=i, column=1, value=date_str).font = _body_font(size=9)
        ws.cell(row=i, column=1).border = _border()

        for name, result in bt_results.items():
            val = result.loc[date, 'Portfolio_Value'] if date in result.index else None
            c = ws.cell(row=i, column=strategy_cols[name], value=round(val, 2) if val else None)
            c.number_format = '$#,##0.00'
            c.font = _body_font(size=9); c.border = _border()

    # Gráfico de línea
    data_rows = i - header_row
    if data_rows > 1:
        chart = LineChart()
        chart.title    = 'Evolución del Portafolio por Estrategia'
        chart.style    = 10
        chart.y_axis.title = 'Valor ($)'
        chart.x_axis.title = 'Período'
        chart.height   = 12; chart.width = 22

        for j, name in enumerate(bt_results.keys(), start=2):
            data_ref = Reference(ws, min_col=j, min_row=header_row,
                                 max_row=header_row + data_rows)
            chart.add_data(data_ref, titles_from_data=True)

        ws.add_chart(chart, f'J{row_offset}')

    _auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 5: Predicciones
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_predictions(wb: Workbook, predictor, metrics: dict):
    ws = wb.create_sheet('Predicciones')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:G1')
    ws['A1'].value     = '🔮 Resultados de Modelos Predictivos'
    ws['A1'].font      = _header_font(size=14)
    ws['A1'].fill      = _fill(DARK_BLUE)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    # Métricas
    ws['A3'].value = 'Métricas de Error por Modelo'
    ws['A3'].font  = _header_font(); ws['A3'].fill = _fill(MID_BLUE)

    for j, h in enumerate(['Modelo', 'MAE', 'RMSE', 'MAPE (%)'], start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = _header_font(size=9); c.fill = _fill(DARK_GRAY)
        c.alignment = _center(); c.border = _border()

    best_mape = min(m['MAPE(%)'] for m in metrics.values())
    for i, (model, m) in enumerate(metrics.items(), start=5):
        row_fill = _fill(LIGHT_BLUE) if m['MAPE(%)'] == best_mape else _fill(WHITE)
        for j, val in enumerate([model, m['MAE'], m['RMSE'], m['MAPE(%)']], start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = _body_font(); c.fill = row_fill
            c.alignment = _center(); c.border = _border()
            if j > 1:
                c.number_format = '0.00'

    # Nota al mejor modelo
    note_row = 5 + len(metrics)
    ws.cell(row=note_row, column=1,
            value=f'✅ Mejor modelo (menor MAPE): {min(metrics, key=lambda k: metrics[k]["MAPE(%)"])}')
    ws.cell(row=note_row, column=1).font = Font(name='Arial', size=10, bold=True, color=GREEN)

    # Predicciones vs real
    header_row = note_row + 2
    ws.cell(row=header_row, column=1, value='Comparativa Predicción vs. Real')
    ws.cell(row=header_row, column=1).font = _header_font()
    ws.cell(row=header_row, column=1).fill = _fill(MID_BLUE)

    col_headers = ['Fecha', 'Precio Real'] + list(predictor.predictions.keys())
    for j, h in enumerate(col_headers, start=1):
        c = ws.cell(row=header_row + 1, column=j, value=h)
        c.font = _header_font(size=9); c.fill = _fill(DARK_GRAY)
        c.alignment = _center(); c.border = _border()

    test_close = predictor.test['Close']
    for i, (date, real_val) in enumerate(test_close.items(), start=header_row + 2):
        date_str = date.strftime('%d/%m/%Y') if hasattr(date, 'strftime') else str(date)
        row_fill = _fill(WHITE) if i % 2 == 0 else _fill(LIGHT_GRAY)

        ws.cell(row=i, column=1, value=date_str).font = _body_font(size=9)
        ws.cell(row=i, column=1).border = _border()

        c = ws.cell(row=i, column=2, value=round(real_val, 2))
        c.number_format = '$#,##0.00'; c.font = _body_font(size=9)
        c.fill = row_fill; c.border = _border()

        for j, (model, preds) in enumerate(predictor.predictions.items(), start=3):
            pred_val = preds.get(date, None)
            c = ws.cell(row=i, column=j,
                        value=round(pred_val, 2) if pred_val is not None else None)
            c.number_format = '$#,##0.00'
            c.font = Font(name='Arial', size=9, color=LINK_GREEN)
            c.fill = row_fill; c.border = _border()

    # Gráfico de líneas
    last_data_row = header_row + 1 + len(test_close)
    chart = LineChart()
    chart.title    = f'Predicciones vs. Precio Real — {predictor.ticker}'
    chart.style    = 10
    chart.y_axis.title = 'Precio ($)'
    chart.height   = 12; chart.width = 22

    n_cols = 2 + len(predictor.predictions)
    data_ref = Reference(ws, min_col=2, max_col=n_cols,
                         min_row=header_row + 1, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    ws.add_chart(chart, f'I{header_row}')

    _auto_col_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Función principal de exportación
# ─────────────────────────────────────────────────────────────────────────────

def export_to_excel(
    kpis: dict,
    portfolio: dict,
    transactions: pd.DataFrame,
    bt_results: dict,
    predictor,
    pred_metrics: dict,
    output_path: str = 'outputs/financial_report.xlsx'
):
    """Genera el reporte Excel completo con todas las hojas."""
    wb = Workbook()

    print("  → Hoja 1: Dashboard KPIs...")
    _sheet_kpi(wb, kpis)

    print("  → Hoja 2: Portafolio de acciones...")
    _sheet_portfolio(wb, portfolio)

    print("  → Hoja 3: Transacciones...")
    _sheet_transactions(wb, transactions)

    print("  → Hoja 4: Backtesting...")
    _sheet_backtest(wb, bt_results)

    print("  → Hoja 5: Predicciones...")
    _sheet_predictions(wb, predictor, pred_metrics)

    wb.save(output_path)
    print(f"  ✅ Reporte guardado en: {output_path}")
    return output_path